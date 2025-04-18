import win32com.client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re

# === 設定 ===
input_path = r"C:\path\to\your\file.xlsx"
output_path = r"C:\path\to\chart_headers_decorated.xlsx"
target_sheet_name = "グラフシート"

# === Excel起動 ===
excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = False
wb = excel.Workbooks.Open(input_path)
ws_dict = {ws.Name: ws for ws in wb.Sheets}

# === 出力用ブック作成 ===
output_wb = Workbook()
output_ws = output_wb.active
output_ws.title = "HeaderInfo"

# ヘッダー定義
headers = [
    "Chart Index (見た目順)", "Series Name",
    "X Sheet", "X Column", "X Header1", "X Header2", "X Header3",
    "Y Sheet", "Y Column", "Y Header1", "Y Header2", "Y Header3",
    "X Axis Label", "Y Axis Label", "TopLeft Cell"
]
output_ws.append(headers)

# === 装飾定義 ===
header_fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")
header_font = Font(bold=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# ヘッダー装飾
for col in range(1, len(headers) + 1):
    cell = output_ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font

# === ヘルパー関数 ===

def parse_formula(formula):
    try:
        inner = formula.strip()[len("=SERIES("):-1]
        parts = [p.strip() for p in inner.split(",")]
        if len(parts) >= 3:
            return parts[1], parts[2]
    except:
        pass
    return None, None

def parse_range_for_header(range_str):
    if not range_str:
        return None, None
    match = re.match(
        r"^(?:'(?P<sheet_quoted>[^']+)'|(?P<sheet_unquoted>[^!]+))!\$?(?P<col>[A-Z]+)\$?\d+:\$?(?P=col)\$?\d+",
        range_str.strip()
    )
    if match:
        sheet_name = match.group("sheet_quoted") or match.group("sheet_unquoted")
        col = match.group("col")
        return sheet_name, col
    return None, None

def get_headers(sheet_name, col_letter):
    if not sheet_name or not col_letter:
        return ["", "", ""]
    try:
        ws = ws_dict.get(sheet_name)
        return [ws.Range(f"{col_letter}{r}").Value for r in range(1, 4)]
    except:
        return ["", "", ""]

def get_axis_title(chart, axis_type):
    try:
        axis = chart.Axes(axis_type)
        if axis.HasTitle:
            return axis.AxisTitle.Text
    except:
        pass
    return ""

# === グラフ情報収集とソート ===

sheet = wb.Sheets(target_sheet_name)
chart_objects = sheet.ChartObjects()

# グラフ情報を取得
chart_info_list = []
for i in range(1, chart_objects.Count + 1):
    chart_obj = chart_objects.Item(i)
    chart = chart_obj.Chart
    top = chart_obj.Top
    left = chart_obj.Left
    try:
        top_left_cell = chart_obj.TopLeftCell.Address.replace("$", "")
    except:
        top_left_cell = ""
    chart_info_list.append({
        "top": top,
        "left": left,
        "chart": chart,
        "top_left_cell": top_left_cell
    })

# 見た目順にソート（左上→右下）
chart_info_list.sort(key=lambda c: (c["top"], c["left"]))

# === 出力 ===

index_ranges = {}  # チャートインデックス → [開始行, 終了行]
current_row = 2  # データ書き込み開始行

for display_index, info in enumerate(chart_info_list, start=1):
    chart = info["chart"]
    top_left_cell = info["top_left_cell"]
    x_axis_label = get_axis_title(chart, 1)
    y_axis_label = get_axis_title(chart, 2)

    for s in chart.SeriesCollection():
        name = str(s.Name)
        formula = str(s.Formula)
        x_range, y_range = parse_formula(formula)

        x_sheet, x_col = parse_range_for_header(x_range)
        y_sheet, y_col = parse_range_for_header(y_range)

        x_headers = get_headers(x_sheet, x_col)
        y_headers = get_headers(y_sheet, y_col)

        row = [
            display_index, name,
            x_sheet or "", x_col or "", *x_headers,
            y_sheet or "", y_col or "", *y_headers,
            x_axis_label, y_axis_label, top_left_cell
        ]
        output_ws.append(row)

        # 行範囲記録
        if display_index not in index_ranges:
            index_ranges[display_index] = [current_row, current_row]
        else:
            index_ranges[display_index][1] = current_row

        current_row += 1

# === チャートインデックスごとにブロック罫線適用 ===
for start_row, end_row in index_ranges.values():
    for row in range(start_row, end_row + 1):
        for col in range(1, len(headers) + 1):
            output_ws.cell(row=row, column=col).border = thin_border

# === 列幅自動調整 ===
for col_cells in output_ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
    col_letter = get_column_letter(col_cells[0].column)
    output_ws.column_dimensions[col_letter].width = max_length + 2

# === 保存と終了 ===
output_wb.save(output_path)
wb.Close(False)
excel.Quit()
print(f"✅ 完了: {output_path}")